from selenium import webdriver
from selenium.webdriver.support.select import Select
from bs4 import BeautifulSoup

import time
import random

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import constants
# from proxies import PROXY
from console_logging import LOGGER


def parse_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('div', class_='climate-table-wrap')
    trs = table.find_all('tr')
    trs = trs[2:]

    month = []
    for tr in trs:
        tds = tr.find_all('td')
        month.append([tds[2].text, tds[5].text])

    return month


def process_data(month_data, month, year):
    month = month if month > 9 else '0' + str(month)
    for index, day_data in enumerate(month_data):
        try:
            day_data = list(map(float, day_data))

            day = index + 1 if index + 1 > 9 else '0' + str(index + 1)

            month_data[index] = [f'{str(day)}.{month}.{year}'] + \
                                [int(num + (0.5 if num > 0 else -0.5)) for num in day_data]
        except:
            LOGGER.log(f'error at day {index + 1}...')


def upload_data_to_excel(month_data):
    # append new month to end of excel and save
    wb = load_workbook(constants.xlsx_name)
    page = wb.active
    for day_data in month_data:
        page.append(day_data)

    wb.save(filename=constants.xlsx_name)


def get_weathers(url):
    profile = webdriver.FirefoxProfile()
    profile.set_preference("general.useragent.override", random.choice(constants.headers))
    browser = webdriver.Firefox(profile, executable_path=constants.gecko_path)

    browser.get(url)

    for year in range(constants.start_year, constants.final_year + 1):
        try:
            select_element = browser.find_element_by_xpath(
                '/html/body/div/main/div/div/div/div/div/div[5]/div[2]/div[2]/form/div/div[2]/select')
            select_object = Select(select_element)
            select_object.select_by_value(str(year))

            for month in range(1, 13):
                if year == constants.final_year and month == constants.final_month:
                    break
                try:
                    select_element = browser.find_element_by_xpath(
                        '/html/body/div/main/div/div/div/div/div/div[5]/div[2]/div[2]/form/div/div[1]/select')
                    select_object = Select(select_element)
                    select_object.select_by_value(str(month))

                    browser.find_element_by_xpath(
                        '/html/body/div/main/div/div/div/div/div/div[5]/div[2]/div[2]/form/div/button').click()

                    LOGGER.log(f'processing data from {year}, {month}')
                    time.sleep(random.uniform(3, 6))

                    month_data = parse_html(browser.page_source)
                    process_data(month_data, month, year)
                    upload_data_to_excel(month_data)
                except:
                    LOGGER.log(f'error at month {month}...')
        except:
            LOGGER.log(f'error at year {year}...')

    browser.quit()


def main():
    # create blank excel
    wb = Workbook()
    page = wb.active
    page.title = 'weather'
    page.append(['date', 'avg', 'precipitation'])
    wb.save(filename=constants.xlsx_name)

    # main logic of parsing and loading data every month
    get_weathers(constants.url_weather)
    LOGGER.log(f'data is ready')


if __name__ == '__main__':
    main()
